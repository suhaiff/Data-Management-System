from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session, jsonify
import os
import pandas as pd
from werkzeug.utils import secure_filename
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
import plotly.express as px
import plotly.io as pio
import uuid
import numbers
import numpy as np

# Configuration
BASE_DIR = os.path.dirname(__file__)
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploaded_files')
EXCEL_FILE = os.path.join(BASE_DIR, 'file_log.xlsx')
SHEET_NAME = 'file header' # exactly as requested
# Set to None to accept any file type. To restrict, set to a set of extensions like {'pdf','png','jpg'}
ALLOWED_EXTENSIONS = None

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = 'change-me-to-a-secure-random-string'

# ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Retry helper (Windows file locks)
def _with_retry(func, retries=5, delay=0.3):
    for i in range(retries):
        try:
            return func()
        except PermissionError:
            if i == retries - 1:
                raise
            time.sleep(delay)

            
def allowed_file(filename):
    if ALLOWED_EXTENSIONS is None:
        return True
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def ensure_excel_sheet():
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if "file header" not in wb.sheetnames:
            ws1 = wb.create_sheet("file header")
            ws1.append(['S.No', 'Path', 'file type', 'file name', 'file upload date', 'file upload time'])
        if "file details" not in wb.sheetnames:
            ws2 = wb.create_sheet("file details")
            ws2.append(['S.No', 'Path', 'file type', 'Sheet name', 'file name', 'file upload date', 'file upload time'])
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "file header"
        ws1.append(['S.No', 'Path', 'file type', 'file name', 'file upload date', 'file upload time'])
        ws2 = wb.create_sheet("file details")
        ws2.append(['S.No', 'Path', 'file type', 'Sheet name', 'file name', 'file upload date', 'file upload time'])
    return wb


@app.route('/')
def home():
    return render_template('index.html', active_page='home')

# @app.route('/upload')
# def upload():
#     return render_template('upload.html', active_page='upload')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        uploaded_file = request.files.get('file')
        if not uploaded_file or uploaded_file.filename == '':
            flash("Please select a file")
            return redirect(url_for('upload'))

        filename = secure_filename(uploaded_file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        saved_filename = f"{timestamp}_{filename}"
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
        uploaded_file.save(save_path)

        # Open or create Excel file
        wb = ensure_excel_sheet()

        # Common values
        rel_path = os.path.relpath(save_path, BASE_DIR)
        file_type = uploaded_file.mimetype
        today = datetime.now().strftime("%Y-%m-%d")
        now_time = datetime.now().strftime("%H:%M:%S")

        # 1️⃣ Log into file header
        ws1 = wb["file header"]
        s_no_header = ws1.max_row  # since first row is header
        ws1.append([s_no_header, rel_path, file_type, filename, today, now_time])

        # 2️⃣ Log into file details (one row per sheet in uploaded Excel)
        try:
            uploaded_wb = load_workbook(save_path, read_only=True)
            ws2 = wb["file details"]
            for sheet in uploaded_wb.sheetnames:
                s_no_detail = ws2.max_row  # since first row is header
                ws2.append([s_no_detail, rel_path, file_type, sheet, filename, today, now_time])
            uploaded_wb.close()
        except Exception as e:
            flash(f"Could not read sheets from uploaded file: {e}")

        # Save and close
        wb.save(EXCEL_FILE)
        wb.close()

        flash("File uploaded successfully")
        return redirect(url_for('configuration'))

    return render_template('upload.html', active_page='upload')


# @app.route('/configuration')
# def configuration():
#     return render_template('configuration.html', active_page='configuration')

@app.route('/configuration', methods=['GET', 'POST'])
def configuration():
    import os
    from openpyxl import Workbook, load_workbook
    from flask import flash, redirect, url_for, request, render_template

    # 1️⃣ Collect all sheet names from uploaded Excel files
    sheet_names = set()  # use set to avoid duplicates
    if os.path.exists(EXCEL_FILE):
        wb_log = load_workbook(EXCEL_FILE, read_only=True)
        if "file details" in wb_log.sheetnames:
            ws_details = wb_log["file details"]
            for row in ws_details.iter_rows(min_row=2, values_only=True):
                sheet_in_file = row[3]  # column 3 = Sheet Name
                if sheet_in_file:
                    sheet_names.add(sheet_in_file)
        wb_log.close()
    sheet_names = sorted(list(sheet_names))  # convert set to sorted list

    if request.method == 'POST':
        # 2️⃣ Get all table names and selected sheets from the form
        table_names = request.form.getlist("tableName[]")
        selected_sheets = request.form.getlist("sheetName[]")

        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()

        # 3️⃣ Create or reset "configuration log" sheet
        config_sheet_name = "configuration log"
        if config_sheet_name in wb.sheetnames:
            ws_config = wb[config_sheet_name]
            wb.remove(ws_config)
        ws_config = wb.create_sheet(config_sheet_name)

        # 4️⃣ Write header
        ws_config.append(["S.No", "Path", "File Name", "Sheet Name"])

        # 5️⃣ Log selected sheets
        if "file details" in wb.sheetnames:
            ws_details = wb["file details"]
            file_rows = list(ws_details.iter_rows(min_row=2, values_only=True))
            for idx, (table_name, sheet_name) in enumerate(zip(table_names, selected_sheets), start=1):
                # Find the corresponding file info from file details
                for row in file_rows:
                    path, file_name, sheet_in_file = row[1], row[4], row[3]
                    if sheet_in_file == sheet_name:
                        ws_config.append([idx, path, file_name, sheet_name])
                        break

        # 6️⃣ Save workbook
        wb.save(EXCEL_FILE)
        wb.close()

        flash("Configuration saved successfully!")
        return redirect(url_for('report_configuration'))

    # 7️⃣ Render template with sheet names for dropdown
    return render_template('configuration.html', active_page='configuration', sheet_names=sheet_names)




@app.route('/report-configuration', methods=['GET', 'POST'])
def report_configuration():
    # Read configuration log (sheets selected in previous step)
    config_rows = []
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if "configuration log" in wb.sheetnames:
            ws = wb["configuration log"]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if any(r):
                    config_rows.append({
                        "sno": r[0],
                        "path": r[1],
                        "file_name": r[2],
                        "sheet_name": r[3]
                    })
        wb.close()

    if request.method == 'POST':
        header_map = request.form.to_dict()   # { "sheetname": "row_number", ... }

        wb = load_workbook(EXCEL_FILE)
        # Create or get file_header_columns sheet
        if "file_header_columns" in wb.sheetnames:
            ws_cols = wb["file_header_columns"]
        else:
            ws_cols = wb.create_sheet("file_header_columns")
            ws_cols.append(["S.No", "Path", "File Name", "Sheet Name", "Column Name", "SheetName_ColumnName"])

        s_no = ws_cols.max_row  # continue numbering

        for cfg in config_rows:
            sheet_name = cfg["sheet_name"]
            header_row = int(header_map.get(sheet_name, 1))  # default row=1
            file_path = os.path.join(BASE_DIR, cfg["path"])
            file_name = cfg["file_name"]

            # open uploaded Excel file
            if os.path.exists(file_path):
                wb_file = load_workbook(file_path, read_only=True)
                if sheet_name in wb_file.sheetnames:
                    ws_file = wb_file[sheet_name]
                    headers = [cell.value for cell in ws_file[header_row] if cell.value]
                    for col in headers:
                        s_no += 1
                        ws_cols.append([
                            s_no,
                            cfg["path"],
                            file_name,
                            sheet_name,
                            col,
                            f"{sheet_name}_{col}"
                        ])
                wb_file.close()

        wb.save(EXCEL_FILE)
        wb.close()
        flash("Report configuration saved. Columns extracted successfully!")
        return redirect(url_for('data_model'))

    return render_template('report_configuration.html', active_page='report_configuration', config_rows=config_rows)

# ==========================
# Data Model Page (Join Builder)
# ==========================
@app.route('/data-model', methods=['GET', 'POST'])
def data_model():
    import json
    from openpyxl import load_workbook
    import pandas as pd

    available_columns, option_map, sheet_to_path = {}, {}, {}

    # 0️⃣ Load metadata from file_header_columns sheet
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
            if "file_header_columns" in wb.sheetnames:
                ws = wb["file_header_columns"]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    # expected row structure: [S.No, Path, File Name, Sheet Name, Column Name, SheetName_ColumnName]
                    path, file_name, sheet, col, opt = r[1], r[2], r[3], r[4], r[5]
                    if not sheet or not col or not opt:
                        continue
                    sheet, col, opt = str(sheet).strip(), str(col).strip(), str(opt).strip()
                    available_columns.setdefault(sheet, []).append({"col": col, "option": opt})
                    option_map[opt] = (sheet, col)
                    if sheet not in sheet_to_path and path:
                        sheet_to_path[sheet] = str(path).replace('/', os.sep).replace('\\', os.sep)
            wb.close()
        except Exception as e:
            flash(f"Error reading file_header_columns: {e}", "danger")
            return render_template('data_model.html',
                                   active_page='data_model',
                                   sheet_names=[],
                                   cols_by_sheet={},
                                   available_columns={})

    if not available_columns:
        available_columns = {"NoSheets": [{"col": "NoColumnsFound", "option": "N/A"}]}

    # Cache for dataframes
    df_cache = {}

    def load_sheet_df(sheet_name):
        """Load sheet as DataFrame and cache it."""
        if sheet_name not in sheet_to_path:
            return None, f"No file path found for '{sheet_name}'."
        path_abs = os.path.join(BASE_DIR, sheet_to_path[sheet_name])
        if not os.path.exists(path_abs):
            return None, f"File not found: {path_abs}"
        key = (path_abs, sheet_name)
        if key in df_cache:
            return df_cache[key], None
        try:
            df = pd.read_excel(path_abs, sheet_name=sheet_name, engine="openpyxl")
            # normalize column names
            df.columns = pd.Index(map(lambda x: str(x).strip(), df.columns))
            df_cache[key] = df
            return df, None
        except Exception as e:
            return None, f"Failed to load {sheet_name}: {e}"

    def _write_with_retry(fn, attempts=3):
        err = None
        for _ in range(attempts):
            try:
                fn()
                return
            except Exception as ex:
                err = ex
                time.sleep(0.2)
        if err:
            raise err

    def _parse_sheet_col(v):
        """Return (sheet, column) either from option_map or from 'Sheet_Column' style value."""
        if not v:
            return None, None
        if v in option_map:
            return option_map[v]
        parts = str(v).split('_', 1)
        return (parts[0], parts[1]) if len(parts) == 2 else (None, None)

    # ==========================================
    # POST: Handle Join Logic
    # ==========================================
    if request.method == 'POST':
        left_opts = request.form.getlist("left_col[]")
        right_opts = request.form.getlist("right_col[]")
        join_types = request.form.getlist("join_type[]")

        bridge_toggle = str(request.form.get("bridge_toggle", "")).lower() in ("on", "true", "1", "yes")
        create_preview = str(request.form.get("create_preview", "")).lower() in ("on", "true", "1", "yes")

        if not left_opts or not right_opts or not join_types:
            flash("Invalid submission: missing join configuration.", "danger")
            return redirect(url_for('data_model'))

        join_map = {"Inner Join": "inner", "Left Join": "left", "Right Join": "right", "Outer Join": "outer"}
        l_opt, r_opt, jt = left_opts[0], right_opts[0], join_types[0]

        if jt not in join_map:
            flash(f"Unsupported join type: {jt}", "danger")
            return redirect(url_for('data_model'))

        l_sheet, l_col = _parse_sheet_col(l_opt)
        r_sheet, r_col = _parse_sheet_col(r_opt)
        if not all([l_sheet, l_col, r_sheet, r_col]):
            flash("Invalid column references selected.", "danger")
            return redirect(url_for('data_model'))

        df_left, err = load_sheet_df(l_sheet)
        if err:
            flash(err, "danger")
            return redirect(url_for('data_model'))
        df_right, err = load_sheet_df(r_sheet)
        if err:
            flash(err, "danger")
            return redirect(url_for('data_model'))

        # Normalize key columns to string to reduce mismatches (preserve NaN -> pd.NA)
        def _norm(s): return s.astype(str).where(~s.isna(), pd.NA)
        df_left[l_col], df_right[r_col] = _norm(df_left[l_col]), _norm(df_right[r_col])

        # Detect cardinality: used only for blocking expand on m:m
        left_dups = df_left.duplicated(l_col).any()
        right_dups = df_right.duplicated(r_col).any()
        cardinality = ("1:1" if not left_dups and not right_dups else
                       "m:1" if left_dups and not right_dups else
                       "1:m" if not left_dups and right_dups else "m:m")

        # ---------------------------
        # BRIDGE MODE: produce aggregated joined_output grouped by the shared key
        # ---------------------------
        if bridge_toggle:
            # We intentionally use the raw column name (e.g., "Sale_ID") as the canonical key
            canonical_key = l_col  # keep same name as left's column (user expects this)
            app.logger.info(f"Bridge Mode Active — Using canonical key column: {canonical_key}")

            # Build unique bridge keys (all unique values from both sides)
            bridge = pd.DataFrame({canonical_key: pd.unique(pd.concat([df_left[l_col], df_right[r_col]], ignore_index=True))})

            # Prepare left aggregation
            left = df_left.copy()
            left[canonical_key] = left[l_col]  # ensure canonical key column exists in left

            # numeric and categorical columns on left (exclude the key)
            left_numeric = [c for c in left.columns if c != canonical_key and pd.api.types.is_numeric_dtype(left[c])]
            left_categorical = [c for c in left.columns if c != canonical_key and c not in left_numeric]

            # group and aggregate (sum numeric, first non-null for categorical)
            if left_numeric:
                left_num_agg = left.groupby(canonical_key, dropna=False)[left_numeric].sum(min_count=1)
            else:
                left_num_agg = pd.DataFrame(index=bridge[canonical_key]).rename_axis(canonical_key)

            if left_categorical:
                def first_non_null(x):
                    nz = x.dropna()
                    return nz.iloc[0] if not nz.empty else ""
                left_cat_agg = left.groupby(canonical_key, dropna=False)[left_categorical].agg(first_non_null)
            else:
                left_cat_agg = pd.DataFrame(index=bridge[canonical_key]).rename_axis(canonical_key)

            left_summary = pd.concat([left_num_agg, left_cat_agg], axis=1).reset_index()

            # Prepare right aggregation similarly
            right = df_right.copy()
            right[canonical_key] = right[r_col]

            right_numeric = [c for c in right.columns if c != canonical_key and pd.api.types.is_numeric_dtype(right[c])]
            right_categorical = [c for c in right.columns if c != canonical_key and c not in right_numeric]

            if right_numeric:
                right_num_agg = right.groupby(canonical_key, dropna=False)[right_numeric].sum(min_count=1)
            else:
                right_num_agg = pd.DataFrame(index=bridge[canonical_key]).rename_axis(canonical_key)

            if right_categorical:
                def first_non_null(x):
                    nz = x.dropna()
                    return nz.iloc[0] if not nz.empty else ""
                right_cat_agg = right.groupby(canonical_key, dropna=False)[right_categorical].agg(first_non_null)
            else:
                right_cat_agg = pd.DataFrame(index=bridge[canonical_key]).rename_axis(canonical_key)

            right_summary = pd.concat([right_num_agg, right_cat_agg], axis=1).reset_index()

            # Merge left & right summaries on canonical_key (outer to keep all bridge values)
            summary = pd.merge(left_summary, right_summary, on=canonical_key, how='outer', suffixes=('_L', '_R'))

            # Normalize column names: for numeric columns create "Sum of <colname>" label
            # Identify numeric columns from left/right to rename
            def _rename_numeric_cols(df_cols):
                rename_map = {}
                for col in df_cols:
                    if col == canonical_key:
                        continue
                    base = col
                    if col.endswith('_L') or col.endswith('_R'):
                        base = col.rsplit('_', 1)[0]
                    # ✅ Use clean column names — no "Sum of"
                    rename_map[col] = base
                return rename_map

            # Build rename map for columns in summary
            rename_map = _rename_numeric_cols(summary.columns.tolist())
            # But be careful if both sides have same numeric (like Total_Amount on right only) we don't duplicate rename,
            # columns already will be 'Total_Amount' (if no suffix) or 'Total_Amount_L'/'Total_Amount_R' with suffix.
            # The logic above maps numeric columns to "Sum of <base>" while keeping categorical base names.

            summary = summary.rename(columns=rename_map)

            # After rename some duplicate column names could still exist: keep first occurrence
            summary = summary.loc[:, ~summary.columns.duplicated()]

            # Ensure numeric columns are numeric and NaN -> 0 for sums
            for c in summary.columns:
                if c == canonical_key:
                    continue
                try:
                    if pd.api.types.is_numeric_dtype(summary[c]):
                        summary[c] = pd.to_numeric(summary[c], errors='coerce').fillna(0)
                except Exception:
                    # if dtype check fails, try convert anyway
                    try:
                        summary[c] = pd.to_numeric(summary[c], errors='coerce').fillna(0)
                    except Exception:
                        pass

            # Final: write Bridge_Key + joined_output (aggregated) + optionally linked sheets
            def _write_bridge_and_joined():
                mode = 'a' if os.path.exists(EXCEL_FILE) else 'w'
                with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode=mode,
                                    if_sheet_exists='replace' if mode == 'a' else None) as writer:
                    bridge.to_excel(writer, sheet_name='Bridge_Key', index=False)
                    # left and right linked raw (for debug/tracing)
                    left.to_excel(writer, sheet_name=f'{l_sheet}_linked', index=False)
                    right.to_excel(writer, sheet_name=f'{r_sheet}_linked', index=False)
                    # aggregated joined output (grouped by canonical key)
                    summary.to_excel(writer, sheet_name='joined_output', index=False)

            _write_with_retry(_write_bridge_and_joined)

            flash("✅ Bridge aggregated: 'joined_output' created (grouped by key, numeric columns summed).", "success")
            return redirect(url_for('data_model_2'))

        # ---------------------------
        # EXPAND MODE (normal join, blocked on m:m)
        # ---------------------------
        else:
            if cardinality == "m:m":
                flash("❌ Expand blocked for many-to-many joins. Use Bridge mode.", "danger")
                return redirect(url_for('data_model'))

            how = join_map[jt]
            result_df = pd.merge(df_left, df_right, left_on=l_col, right_on=r_col,
                                 how=how, suffixes=('_L', '_R'))
            # remove duplicated columns that came from suffixing where possible
            result_df = result_df.loc[:, ~result_df.columns.duplicated()]

            def _write_expand():
                mode = 'a' if os.path.exists(EXCEL_FILE) else 'w'
                with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode=mode,
                                    if_sheet_exists='replace' if mode == 'a' else None) as writer:
                    result_df.to_excel(writer, sheet_name='joined_output', index=False)

            _write_with_retry(_write_expand)
            flash("✅ Expand completed: 'joined_output' created.", "success")
            return redirect(url_for('data_model_2'))

    # ==========================================
    # GET: Load Join Builder UI
    # ==========================================
    sheet_names = list(available_columns.keys())
    cols_by_sheet = {s: [c["col"] for c in cols] for s, cols in available_columns.items()}

    return render_template('data_model.html',
                           active_page='data_model',
                           sheet_names=sheet_names,
                           cols_by_sheet=cols_by_sheet,
                           available_columns=available_columns)
# # Data Model page (join builder) - accepts both /data-model and /data_model
# @app.route('/data-model', methods=['GET', 'POST'])
# @app.route('/data_model', methods=['GET', 'POST'])
# def data_model():
#     # Load available columns from file_header_columns sheet
#     available_columns = {}      # { "SheetName": [ {"col": column_name, "option": option_value}, ... ] }
#     option_map = {}             # { "SheetName_ColumnName": (sheet, column) }
#     sheet_to_path = {}          # { "SheetName": relative_path }

#     if os.path.exists(EXCEL_FILE):
#         try:
#             wb = load_workbook(EXCEL_FILE, read_only=True)
#             if "file_header_columns" in wb.sheetnames:
#                 ws = wb["file_header_columns"]
#                 for r in ws.iter_rows(min_row=2, values_only=True):
#                     # expected: [S.No, Path, File Name, Sheet Name, Column Name, SheetName_ColumnName]
#                     path = r[1]; file_name = r[2]; sheet = r[3]; col = r[4]; opt = r[5]
#                     if sheet and opt:
#                         available_columns.setdefault(sheet, []).append({"col": col, "option": opt})
#                         option_map[opt] = (sheet, col)
#                         # track the path for that sheet (first occurrence)
#                         if sheet not in sheet_to_path and path:
#                             sheet_to_path[sheet] = path
#             wb.close()
#         except Exception as e:
#             flash(f"Error reading file_header_columns: {e}")
#             return render_template('data_model.html', active_page='data_model',
#                                    available_columns=available_columns)

    
        
#     # ✅ Simple Many-to-Many detector
#     def is_many_to_many(df_left, df_right, left_col, right_col):
#         return df_left[left_col].duplicated().any() and df_right[right_col].duplicated().any()    

#     if request.method == 'POST':
#         left_opts = request.form.getlist("left_col[]")
#         right_opts = request.form.getlist("right_col[]")
#         join_types = request.form.getlist("join_type[]")

#         # Basic validation
#         if not (left_opts and right_opts and join_types) or not (len(left_opts) == len(right_opts) == len(join_types)):
#             flash("Invalid form submission: ensure each row has left, right and join type.")
#             return redirect(url_for('data_model'))

#         # Prepare dataframes cache per (path, sheet) so we don't reload repeatedly
#         df_cache = {}  # key = (path, sheet) -> DataFrame

#         def load_sheet_df(sheet_name):
#             """Load the dataframe for that sheet using sheet_to_path mapping."""
#             if sheet_name not in sheet_to_path:
#                 return None, f"No file path found for sheet '{sheet_name}' in file_header_columns."
#             path_rel = sheet_to_path[sheet_name]
#             path_abs = os.path.join(BASE_DIR, path_rel)
#             if not os.path.exists(path_abs):
#                 return None, f"File not found: {path_abs}"
#             key = (path_abs, sheet_name)
#             if key in df_cache:
#                 return df_cache[key], None
#             try:
#                 df = pd.read_excel(path_abs, sheet_name=sheet_name, engine="openpyxl")
#                 df_cache[key] = df
#                 return df, None
#             except Exception as e:
#                 return None, f"Failed to read {sheet_name} from {path_abs}: {e}"

#         # perform sequential joins (chain must be connected)
#         result_df = None
#         included_sheets = set()
#         join_map = {"Inner Join": "inner", "Left Join": "left", "Right Join": "right", "Outer Join": "outer"}

#         join_log_rows = []  # to record what was joined for the data_model sheet

#         # ✅ Apply Many-to-Many check before merge
#         if is_many_to_many(df_left, df_right, l_col, r_col):
#             flash("⚠️ Many to Many Entries found between selected join columns. Please remove duplicates.")
#             return redirect(url_for('data_model'))

#         try:
#             for l_opt, r_opt, jt in zip(left_opts, right_opts, join_types):
#                 if l_opt not in option_map or r_opt not in option_map:
#                     flash(f"Column mapping not found for {l_opt} or {r_opt}")
#                     return redirect(url_for('data_model'))

#                 l_sheet, l_col = option_map[l_opt]
#                 r_sheet, r_col = option_map[r_opt]

#                 df_left, err = load_sheet_df(l_sheet)
#                 if err:
#                     flash(err); return redirect(url_for('data_model'))
#                 df_right, err = load_sheet_df(r_sheet)
#                 if err:
#                     flash(err); return redirect(url_for('data_model'))

#                 how = join_map.get(jt, "inner")

#                 # ✅ Apply Many-to-Many check before merge
#                 if is_many_to_many(df_left, df_right, l_col, r_col):
#                     flash("⚠️ Many to Many Entries found between selected join columns. Please remove duplicates.")
#                     return redirect(url_for('data_model'))

#                 if result_df is None:
#                     # first join
#                     if l_col not in df_left.columns:
#                         flash(f"Column '{l_col}' not in sheet '{l_sheet}'"); return redirect(url_for('data_model'))
#                     if r_col not in df_right.columns:
#                         flash(f"Column '{r_col}' not in sheet '{r_sheet}'"); return redirect(url_for('data_model'))
#                     result_df = pd.merge(df_left, df_right, left_on=l_col, right_on=r_col, how=how, suffixes=('_L','_R'))
#                     included_sheets.update([l_sheet, r_sheet])
#                 else:
#                     # subsequent join: determine which side (left/right) is already in result_df
#                     # we check presence of columns named exactly l_col or r_col in result_df
#                     if l_col in result_df.columns:
#                         # result_df has left column -> merge with df_right using r_col
#                         if r_col not in df_right.columns:
#                             flash(f"Column '{r_col}' not found in {r_sheet}"); return redirect(url_for('data_model'))
#                         result_df = pd.merge(result_df, df_right, left_on=l_col, right_on=r_col, how=how, suffixes=('','_R2'))
#                         included_sheets.add(r_sheet)
#                     elif r_col in result_df.columns:
#                         # result_df has right column -> merge with df_left
#                         if l_col not in df_left.columns:
#                             flash(f"Column '{l_col}' not found in {l_sheet}"); return redirect(url_for('data_model'))
#                         result_df = pd.merge(result_df, df_left, left_on=r_col, right_on=l_col, how=how, suffixes=('','_L2'))
#                         included_sheets.add(l_sheet)
#                     else:
#                         # not connected -> abort with clear message
#                         flash(f"Join chain is not connected for {l_opt} <> {r_opt}. Make sure joins form a connected chain.")
#                         return redirect(url_for('data_model'))

#                 join_log_rows.append((l_opt, r_opt, jt))

#         except Exception as e:
#             flash(f"Unexpected error during join: {e}")
#             return redirect(url_for('data_model'))

#         # If nothing was joined
#         if result_df is None:
#             flash("No joins performed.")
#             return redirect(url_for('data_model'))
        
#         # Clean NaN and NaT values before saving
#         result_df = result_df.fillna("")  # replaces NaN, NaT, None with empty string


#         # Save joined result to EXCEL_FILE as 'joined_output' (replace if exists)
#         try:
#             if os.path.exists(EXCEL_FILE):
#                 with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#                     result_df.to_excel(writer, sheet_name='joined_output', index=False)
#             else:
#                 with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
#                     result_df.to_excel(writer, sheet_name='joined_output', index=False)
#         except Exception as e:
#             flash(f"Failed to write joined_output sheet: {e}")
#             return redirect(url_for('data_model'))

#         # Log the join definitions into 'data_model' sheet
#         try:
#             wb_log = load_workbook(EXCEL_FILE)
#             if "data_model" not in wb_log.sheetnames:
#                 ws_dm = wb_log.create_sheet("data_model")
#                 ws_dm.append(["S.No", "Sheet1_Column", "Sheet2_Column", "Join_Type"])
#             else:
#                 ws_dm = wb_log["data_model"]
#             start_idx = ws_dm.max_row
#             for i, (lopt, ropt, jt) in enumerate(join_log_rows, start=start_idx):
#                 ws_dm.append([i, lopt, ropt, jt])
#             wb_log.save(EXCEL_FILE)
#             wb_log.close()
#         except Exception as e:
#             flash(f"Failed to log data_model details: {e}")
#             return redirect(url_for('data_model'))

#         # Save the joined DataFrame temporarily as JSON in session or just redirect to preview route
#         # We'll just redirect to preview route; preview route will read 'joined_output' sheet.
#         flash("Join completed and saved to 'joined_output'. Redirecting to preview...")
#         return redirect(url_for('data_model_2'))

#     # GET: render builder UI
#     return render_template('data_model.html', active_page='data_model', available_columns=available_columns)



# --- page 2: preview joined output ---
@app.route("/data_model_2", methods=["GET", "POST"])
def data_model_2():
    joined_sheet = "joined_output"

    if not os.path.exists(EXCEL_FILE):
        flash("Excel file not found. Please upload again.")
        return redirect(url_for("data_model"))

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=joined_sheet)
        df_raw = df.copy()
    except Exception as e:
        flash(f"Error loading joined_output: {e}")
        return redirect(url_for("data_model"))

    columns = list(df.columns)

    if request.method == "POST":

        print("Selected columns raw:", request.form.get("selected_columns"))


        # -----------------------------
        # Step 1: Collect all chart configurations
        # -----------------------------
        charts_data = []
        index = 0
        while f"chart_title_{index}" in request.form:
            chart_title = request.form.get(f"chart_title_{index}")
            chart_type = request.form.get(f"chart_type_{index}")
            x_col = request.form.get(f"x_column_{index}")
            y_col = request.form.get(f"y_column_{index}")
            operation = request.form.get(f"operation_{index}")
            font = request.form.get(f"font_{index}")
            color = request.form.get(f"color_{index}")
            width = int(request.form.get(f"width_{index}", 800))
            height = int(request.form.get(f"height_{index}", 600))

            charts_data.append({
                "chart_title": chart_title,
                "chart_type": chart_type,
                "x_col": x_col,
                "y_col": y_col,
                "operation": operation,
                "font": font,
                "color": color,
                "width": width,
                "height": height
            })
            index += 1

            print("Chart config:", chart_type, font, color, width, height)

        # -----------------------------
        # Step 2: Generate each chart
        # -----------------------------
        generated_files = []
        custom_colors = [
            "#E57373", "#64B5F6", "#4DB6AC", "#FFD54F", "#81C784", "#FFB74D", "#9575CD"
        ]
        operation_map = {
            "none": "",
            "sum": "Sum of",
            "avg": "Average of",
            "count": "Count of",
            "distinct_count": "Distinct Count of"
        }

        
        for chart_conf in charts_data:
            chart_title = chart_conf["chart_title"]
            chart_type = chart_conf["chart_type"]
            x_col = chart_conf["x_col"]
            y_col = chart_conf["y_col"]
            operation = chart_conf["operation"]
            font = chart_conf["font"]
            color = chart_conf["color"]
            width = chart_conf["width"]
            height = chart_conf["height"]

            operation_text = operation_map.get(operation, "").strip()
            custom_title = chart_conf.get("chart_title", "").strip()

            if custom_title:
                chart_title = custom_title
            else:
                if chart_type == "card":
                    chart_title = f"Summary of {y_col}"
                elif chart_type == "pie":
                    chart_title = f"{y_col} Distribution by {x_col}"
                elif chart_type == "table":
                    chart_title = f"Data Table"
                elif chart_type == "matrix_table":
                    chart_title = f"Data Table - Matrix(Summable)"
                else:
                    if operation_text:
                        chart_title = f"{operation_text} {y_col} by {x_col}"
                    else:
                        chart_title = f"{y_col} by {x_col}"

            # ---------- Aggregation ----------
            if operation == "sum":
                df_result = df.groupby(x_col)[y_col].sum().reset_index()
            elif operation == "avg":
                df_result = df.groupby(x_col)[y_col].mean().reset_index()
            elif operation == "count":
                df_result = df.groupby(x_col)[y_col].count().reset_index()
                df_result.rename(columns={y_col: f"Count of {y_col}"}, inplace=True)
            elif operation == "distinct_count":
                df_result = df.groupby(x_col)[y_col].nunique().reset_index()
                df_result.rename(columns={y_col: f"Distinct Count of {y_col}"}, inplace=True)
            else:
                df_result = df[[x_col, y_col]]

            y_col_actual = df_result.columns[1]
            # Check if we should show currency
            use_currency = operation in ["sum", "avg"]

            # ---------- Create Chart ----------
            if chart_type == "bar":
                fig = px.bar(
                    df_result,
                    x=x_col,
                    y=y_col_actual,
                    color=x_col,
                    text=df_result[y_col_actual].apply(lambda v: f"₹{v:,.0f}" if use_currency else f"{v}"),
                    title=chart_title,
                    color_discrete_sequence=custom_colors
                )
                fig.update_traces(
                    texttemplate='%{text}',
                    textposition='outside',
                    textangle=90,
                    cliponaxis=False,
                    hovertemplate=f'X: %{{x}}<br>Y: {"₹%{y:,.0f}" if use_currency else "%{y}"}<extra></extra>'
                )
                # Dynamic y-axis with extra space
                max_value = df_result[y_col_actual].max()
                fig.update_yaxes(
                    range=[0, max_value*1.2], dtick=5000
                )
                chart_title = f"<b>{chart_title}</b>"
                fig.update_layout(
                    title=dict(
                        text=chart_title,
                        font=dict(size=25, color=color, family=font),
                        x=0.5,
                        xanchor='center'
                    ),
                    width=width,
                    height=height,
                    uniformtext_minsize=12,
                    uniformtext_mode='show',
                    font=dict(family=font, size=10),
                    title_font_color=color,
                    xaxis_title_font_color=color,
                    yaxis_title_font_color=color,
                    bargap=0.3,
                    bargroupgap=0.1,
                    margin=dict(l=50, r=50, t=80, b=80)
                )
            elif chart_type == "line":
                fig = px.line(
                    df_result,
                    x=x_col,
                    y=y_col_actual,
                    title=chart_title,
                    markers=True
                )
                fig.update_traces(
                    texttemplate=f'{"₹%{y:,.0f}" if use_currency else "%{y}"}',
                    textposition='top center',
                    hovertemplate=f'X: %{{x}}<br>Y: {"₹%{y:,.0f}" if use_currency else "%{y}"}',
                    mode='lines+markers+text'
                )
                chart_title = f"<b>{chart_title}</b>"
                fig.update_layout(
                    title=dict(
                        text=chart_title,
                        font=dict(size=25, color=color, family=font),
                        x=0.5,
                        xanchor='center'
                    ),
                    width=width,
                    height=height, 
                    font=dict(family=font, size=10), 
                    title_font_color=color,
                    xaxis_title_font_color=color,
                    yaxis_title_font_color=color,
                    uniformtext_minsize=12,
                    uniformtext_mode='show',
                    margin=dict(l=50, r=50, t=80, b=80)
                )

            elif chart_type == "pie":
                percentages = df_result[y_col_actual] / df_result[y_col_actual].sum() * 100
                text_positions = ['inside' if p >= 5 else 'outside' for p in percentages]
                pull_values = [0.05 if p < 5 else 0.05 for p in percentages]
                text_templates = [
                    f"%{{label}}: {'₹' if use_currency else ''}%{{value:,.0f}} ({p:.1f}%)" if p >= 5 else f"{'₹' if use_currency else ''}%{{value:,.0f}} ({p:.1f}%)"
                    for p in percentages
                ]
                fig = px.pie(
                    df_result,
                    names=x_col,
                    values=y_col_actual,
                    title=chart_title,
                    color_discrete_sequence=custom_colors
                )
                fig.update_traces(
                    textposition=text_positions,
                    texttemplate=text_templates,
                    insidetextorientation='radial',
                    textfont_size=10,
                    pull=pull_values,
                    marker_line_width=1,
                    hovertemplate=f"%{{label}}: {'₹' if use_currency else ''}%{{value:,.0f}} (%{{percent}})<extra></extra>"
                )
                chart_title = f"<b>{chart_title}</b>"
                fig.update_layout(
                    title=dict(
                        text=chart_title,
                        font=dict(size=25, color=color, family=font),
                        x=0.5,
                        xanchor='center'
                    ),
                    width=width, 
                    height=height, 
                    font=dict(family=font, size=10), 
                    title_font_color=color,
                    margin=dict(t=50,b=50,l=50,r=50), 
                    showlegend=False
                    )
            elif chart_type == "card":
                total_value = df_result[y_col_actual].sum() if operation in ["sum", "none"] else df_result[y_col_actual].mean()
                card_html = f"""
                <div style='
                    display:flex;
                    flex-direction:column;
                    align-items:center;
                    justify-content:center;
                    height:300px;
                    background:linear-gradient(135deg, {color}, #f8f9fa);
                    border-radius:16px;
                    font-family:{font};
                    box-shadow:0 4px 10px rgba(0,0,0,0.1);
                '>
                    <h2 style='color:#333; margin-bottom:10px;'>{chart_title}</h2>
                    <h1 style='font-size:3rem; color:{color}; margin:0;'>₹{total_value:,.0f}</h1>
                </div>
                """

            elif chart_type == "table":
                selected_cols_str = request.form.get(f"selected_columns_{charts_data.index(chart_conf)}", "")
                selected_cols = [col.strip() for col in selected_cols_str.split(",") if col.strip()]

                # Filter only columns that exist in df_result
                existing_cols = [col for col in selected_cols if col in df.columns]
                missing_cols = [col for col in selected_cols if col not in df.columns]

                if missing_cols:
                    print(f"Warning: These columns are missing from df_result: {missing_cols}")

                if not existing_cols:
                    card_html = "<p style='color:red;'>⚠️ No valid columns selected for table.</p>"
                else:
                    df_table = df[existing_cols]

                    # Add serial numbers (1, 2, 3, ...)
                    df_table.insert(0, "S.No", range(1, len(df_table) + 1))

                    # ---------- Format currency columns ----------
                    amount_keywords = ["amount", "total", "transaction", "price", "cost"]
                    currency_cols = [
                        col for col in df_table.columns 
                        if any(keyword.lower() in col.lower() for keyword in amount_keywords)
                    ]
                    for col in currency_cols:
                        df_table[col] = df_table[col].apply(
                            lambda x: f"₹{int(x):,}" if pd.notnull(x) else ""
                            )


                    # Wrap table in a container div and center it
                    table_html = f"""
                    <div style="overflow-x:auto; width:100%; text-align:center; margin:auto;">
                        {df_table.to_html(
                            classes="table table-striped table-bordered table-sm text-center",
                            index=False,
                            escape=False
                        )}
                    </div>
                    """
                    table_html = f"""
                    <style>
                        table th, table td {{
                            text-align: center;
                        }}
                    </style>
                    {table_html}
                    """

                    card_html = f"""
                    <div style='
                        background:#fff;
                        border-radius:12px;
                        box-shadow:0 4px 8px rgba(0,0,0,0.1);
                        overflow:auto;
                        padding:15px;
                    '>
                    <h4 style='
                        font-family:{font}; 
                        color:{color}; 
                        font-size:25px; 
                        text-align:center;
                        font-weight:bold;
                        margin-bottom:22px;
                        '>{chart_title}</h4>
                        {table_html}
                    </div>
                    """
                    

            elif chart_type == "matrix_table":

                # ---- Get user-selected pivot dimensions ----
                row_dim = request.form.get("matrix_row_dim")
                col_dim = request.form.get("matrix_col_dim")
                value_dim = request.form.get("matrix_value_dim")

                # ---- Validate that these exist in dataframe ----
                if not all([row_dim, col_dim, value_dim]):
                    card_html = "<p style='color:red;'>⚠️ Please select Row, Column, and Value fields for matrix table.</p>"
                elif not all(col in df.columns for col in [row_dim, col_dim, value_dim]):
                    card_html = "<p style='color:red;'>⚠️ One or more selected dimensions not found in data.</p>"
                else:
                    # ---- Create Pivot (Matrix) Table ----
                    try:
                        # ---- Create Pivot (Matrix) Table ----
                        df_pivot = pd.pivot_table(
                            df,
                            index=[row_dim],
                            columns=[col_dim],
                            values=value_dim,
                            aggfunc="sum",
                            fill_value=0,
                            margins=True,
                            margins_name="Total"
                        )

                        df_pivot.columns.name = None


                        # Remove extra MultiIndex structure if present
                        if isinstance(df_pivot.columns, pd.MultiIndex):
                            df_pivot.columns = [' '.join(map(str, col)).strip() for col in df_pivot.columns]

                        # Convert pivot index to column (without duplicating)
                        df_pivot.index.name = row_dim
                        df_pivot = df_pivot.reset_index()


                        # ---- Detect whether the selected value column represents currency ----
                        currency_keywords = ["amount", "price", "cost", "value", "total", "sales", "revenue"]
                        is_currency = any(keyword in value_dim.lower() for keyword in currency_keywords)

                        # ---- Format numeric columns ----
                        for col in df_pivot.columns:
                            # If currency type, add ₹ for all numeric cells
                            if is_currency:
                                df_pivot[col] = df_pivot[col].apply(
                                    lambda x: f"₹{int(x):,}" 
                                    if isinstance(x, (int, float)) and not pd.isna(x) and x != 0 
                                    else (f"{int(x):,}" if isinstance(x, (int, float)) and x == 0 else x)
                                )
                            else:
                                # Non-currency (like count) → just keep as number
                                df_pivot[col] = df_pivot[col].apply(
                                    lambda x: f"{int(x):,}" if isinstance(x, (int, float)) and not pd.isna(x) else x
                                )

                        # ---- Generate clean HTML ----
                        pivot_html = df_pivot.to_html(
                            classes="table table-striped table-bordered table-sm text-center",
                            index=False,
                            escape=False
                        )

                        # ---- Apply CSS styling ----
                        table_html = f"""
                        <style>
                            table {{
                                border-collapse: collapse;
                                margin: auto;
                            }}
                            table th, table td {{
                                text-align: center;
                                padding: 6px 10px;
                            }}
                            table tbody tr:last-child {{
                                font-weight: bold;
                                background-color: #f0f0f0;
                            }}
                        </style>
                        <div style="overflow-x:auto; width:100%; text-align:center; margin:auto;">
                            {pivot_html}
                        </div>
                        """

                        # ---- Wrap inside card ----
                        card_html = f"""
                        <div style='
                            background:#fff;
                            border-radius:12px;
                            box-shadow:0 4px 8px rgba(0,0,0,0.1);
                            overflow:auto;
                            padding:15px;
                        '>
                            <h4 style='
                                font-family:{font}; 
                                color:{color}; 
                                font-size:25px; 
                                text-align:center;
                                font-weight:bold;
                                margin-top:14px;
                                margin-bottom:22px;
                            '>
                            {chart_title}</h4>
                            {table_html}
                        </div>
                        """

                    except Exception as e:
                        card_html = f"<p style='color:red;'>⚠️ Error generating matrix table: {e}</p>"

            else:
                continue  # skip unknown chart type



            # ---------- Save chart ----------
            static_dir = os.path.join(BASE_DIR, "static")
            os.makedirs(static_dir, exist_ok=True)
            new_filename = f"chart_{uuid.uuid4().hex}.html"
            new_path = os.path.join(static_dir, new_filename)
            

            if chart_type in ["card", "table", "matrix_table"]:
                chart_id = uuid.uuid4().hex
                full_html = """
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
                    <style>
                        body {{
                            margin: 0;
                            padding: 10px;
                            background: #fafafa;
                        }}
                    </style>
                </head>
                <body>
                    <div class="chart-wrapper" id="chart-%s" style="margin-bottom:20px;">
                        %s
                    </div>
                    <script>
                    window.addEventListener("message", (event) => {
                        if (event.data.type === "updateFilters") {
                            const filters = event.data.filters;
                            console.log("Chart received filters:", filters);
                            
                            const table = document.querySelector("table");
                            if (!table) return;

                            // Get header names from <thead>
                            const headers = Array.from(table.querySelectorAll("thead th")).map(th => th.innerText.trim());
                            const rows = table.querySelectorAll("tbody tr");

                            rows.forEach(row => {
                                let show = true;
                                const cells = Array.from(row.querySelectorAll("td"));

                                // Build a mapping of {columnName: cellText}
                                const rowData = {};
                                headers.forEach((col, i) => {
                                    rowData[col] = cells[i] ? cells[i].innerText.trim() : "";
                                });

                                // Check all active filters
                                for (const [col, values] of Object.entries(filters)) {
                                    if (values.length && !values.includes(rowData[col])) {
                                        show = false;
                                        break;
                                    }
                                }
                                // Show or hide the row
                                row.style.display = show ? "" : "none";
                            });
                        }
                    });
                    </script>
                </body>
                </html>""" % (chart_id, card_html)
                    
                with open(new_path, "w", encoding="utf-8") as f:
                    f.write(full_html)
            else:
                for trace in fig.data:
                    try:
                        trace_name = getattr(trace, "name", None)
                        tx = list(getattr(trace, "x", []))
                        per_point_meta = []

                        # If the trace has bars for categories, attach matching rows from df
                        for xv in tx:
                            # Find all rows matching that category (case-insensitive)
                            matches = df_raw[df_raw[x_col].astype(str).str.lower() == str(xv).lower()]
                            if not matches.empty:
                                # Attach full row info for that category
                                recs = matches.to_dict(orient="records")
                                per_point_meta.append({"records": recs})  # use first match or aggregate
                            else:
                                per_point_meta.append({"records": []})

                        trace.customdata = per_point_meta
                        trace.meta = {"y_field": y_col_actual}
                        print(trace.meta)
                        print(f"✅ Trace '{trace_name}' attached {len(per_point_meta)} customdata entries (x_len={len(tx)})")
                        print(f"   ↳ Sample customdata: {per_point_meta[0] if per_point_meta else 'none'}")
                    except Exception as e:
                        print(f"⚠️ Could not attach customdata for a trace: {e}")

                html_str = pio.to_html(
                    fig,
                    include_plotlyjs="cdn",
                    full_html=True,
                    config={'responsive': True}
                )

                # --- Ensure DOCTYPE at top ---
                if not html_str.lstrip().startswith("<!DOCTYPE html>"):
                    html_str = "<!DOCTYPE html>\n" + html_str

                with open(new_path, "w", encoding="utf-8") as f:
                    f.write(html_str)

                # --- Ensure DOCTYPE at top to prevent Quirks Mode ---
                with open(new_path, "r+", encoding="utf-8") as f:
                    content = f.read()
                    if not content.lstrip().startswith("<!DOCTYPE html>"):
                        f.seek(0)
                        f.write("<!DOCTYPE html>\n" + content)
                        f.truncate()

                # --- Add filter listener inside each chart HTML ---
                filter_listener = """
                <script>
                window.addEventListener("message", (event) => {
                    if (event.data.type === "updateFilters") {
                        const filters = event.data.filters;
                        console.log("Plotly Chart received filters:", filters);

                        
                        const plot = document.querySelector(".plotly-graph-div");
                        if (!plot) return; 

                        if (!plot._originalData) {
                            plot._originalData = JSON.parse(JSON.stringify(plot.data)); // deep copy of traces
                        }

                        const originalData = plot._originalData;

                        originalData.forEach((trace, tIndex) => {
                            console.log(`Trace ${tIndex}:`, {
                                x: trace.x.length,
                                y: trace.y.length,
                                customdata: trace.customdata ? trace.customdata.slice(0, 3) : "no customdata"
                            });
                        });
                        if (Object.values(filters).every(arr => arr.length === 0)) {
                            Plotly.react(plot, originalData, plot.layout);
                            return;
                        }
                        const filteredData = originalData.map(trace => {
                            // if no x/y or no customdata, return trace unchanged
                            if ((!trace.x || !trace.y) && !trace.labels) {
                                return trace;
                            }

                            // deep copy the trace so we don't corrupt originalData
                            const newTrace = JSON.parse(JSON.stringify(trace));
                            

                            const x_src = trace.x || trace.labels || [];
                            const y_src = trace.y || trace.values || [];
                            const customdata = trace.customdata || [];
                            const new_x = [];
                            const new_y = [];
                            const new_cd = [];

                            for (let i = 0; i < x_src.length; i++) {
                                const metaWrapper = customdata[i] || {};
                                const records = metaWrapper.records || [];
                                let include = false;

                                if (records.length === 0) {
                                    include = true; // if no metadata, keep it
                                } else {
                                    // check if ANY record for this bar matches all filters
                                    for (const rec of records) {
                                        let recordMatches = true;
                                        for (const [col, values] of Object.entries(filters)) {
                                            if (!values || values.length === 0) continue;
                                            const recVal = String(rec[col] ?? "").trim().toLowerCase();
                                            const match = values.some(v => String(v).trim().toLowerCase() === recVal);
                                            if (!match) { recordMatches = false; break; }
                                        }
                                        if (recordMatches) {
                                            include = true;
                                            break;
                                        }
                                    }
                                }

                                if (include) {
                                    if (trace.x) new_x.push(trace.x[i]);
                                    // Normalize customdata structure
                                    let records = [];
                                    const cd = trace.customdata?.[i];
                                    if (Array.isArray(cd?.records)) {
                                        records = cd.records;
                                    } else if (cd && typeof cd === "object" && !Array.isArray(cd)) {
                                        records = [cd]; // single flat object
                                    } else {
                                        records = [];
                                    }

                                    console.log(`📊 Records for x: ${trace.x ? trace.x[i] : trace.name}`, records);
                                    console.log(trace.meta);
                                    // Try to detect the numeric field automatically
                                    let totalSales = 0;
                                    for (const rec of records) {
                                        let recordMatches = true;
                                        for (const [col, values] of Object.entries(filters)) {
                                            if (!values || values.length === 0) continue;
                                            const recVal = String(rec[col] ?? "").trim().toLowerCase();
                                            const match = values.some(v => String(v).trim().toLowerCase() === recVal);
                                            if (!match) { recordMatches = false; break; }
                                        }

                                        if (recordMatches) {
                                            // dynamically detect numeric field
                                            const numericField = 
                                                trace.meta?.y_field ||
                                                trace.y_field ||
                                                trace.y_name ||
                                                "Amount"; // fallback if not defined

                                            if (numericField) {
                                                totalSales += Number(rec[numericField]);
                                                console.log(`✅ Matched ${numericField}:`, rec[numericField]);
                                            } else {
                                                console.warn("⚠️ No numeric field found in", rec);
                                            }
                                        }
                                    }

                                    new_y.push(totalSales > 0 ? totalSales : 0);
                                    new_cd.push(trace.customdata[i]);
                                }
                            }

                            if (trace.x) newTrace.x = new_x;
                            if (trace.y) newTrace.y = new_y;

                            // for pie charts
                            if (trace.labels && trace.values) {
                                const new_labels = [];
                                const new_values = [];
                                const new_cd_pie = [];

                                for (let i = 0; i < trace.labels.length; i++) {
                                    const metaWrapper = trace.customdata?.[i] || {};
                                    const records = metaWrapper.records || [];
                                    let include = false;

                                    if (records.length === 0) {
                                        include = true;
                                    } else {
                                        for (const rec of records) {
                                            let recordMatches = true;
                                            for (const [col, values] of Object.entries(filters)) {
                                                if (!values || values.length === 0) continue;
                                                const recVal = String(rec[col] ?? "").trim().toLowerCase();
                                                const match = values.some(v => String(v).trim().toLowerCase() === recVal);
                                                if (!match) { recordMatches = false; break; }
                                            }
                                            if (recordMatches) { include = true; break; }
                                        }
                                    }

                                    if (include) {
                                        new_labels.push(trace.labels[i]);
                                        new_values.push(trace.values[i]);
                                        new_cd_pie.push(metaWrapper);
                                    }
                                }

                                newTrace.labels = new_labels;
                                newTrace.values = new_values;
                                newTrace.customdata = new_cd_pie;
                            }
                            console.log(`✅ Trace '${trace.name}' filtered: kept ${new_x.length} points`, { new_x, new_y });
                            return {
                                ...trace,
                                x: new_x,
                                y: new_y,
                                customdata: new_cd
                            };
                        });

                        // --- 4️⃣ Update the Plotly chart dynamically ---
                        Plotly.react(plot, filteredData, plot.layout);

                    }
                });
                </script>
                """

                # Append the listener before </body> in the generated HTML
                with open(new_path, "r+", encoding="utf-8") as f:
                    html = f.read()
                    if "</body>" in html:
                        html = html.replace("</body>", filter_listener + "\n</body>")
                    else:
                        html += filter_listener
                    f.seek(0)
                    f.write(html)
                    f.truncate()

            generated_files.append(new_filename)

        # -----------------------------
        # Step 3: Store generated files in session
        # -----------------------------
        session["chart_files"] = generated_files
        session["charts_meta"] = charts_data
        print(f"{len(generated_files)} chart(s) generated successfully.")
        return redirect(url_for("chart_view"))

    return render_template("data_model_2.html", columns=columns)



@app.route("/chart_view", methods=["GET"])
def chart_view():
    joined_sheet = "joined_output"

    # load DF (same way you already do)
    if not os.path.exists(EXCEL_FILE):
        flash("Excel file not found.")
        return redirect(url_for("data_model"))

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=joined_sheet)
    except Exception as e:
        flash(f"Error loading joined_output: {e}")
        return redirect(url_for("data_model"))

    
    

    # ----------- Clean NaN, NaT, None values -----------
    df_cleaned = df.copy()

    # Replace NaN/NaT/pd.NA with empty string
    df_cleaned = df_cleaned.replace({pd.NA: "", np.nan: "", "NaT": "", pd.NaT: ""})
    
    # Handle datetime columns cleanly
    for col in df_cleaned.select_dtypes(include=["datetime", "datetimetz"]).columns:
        df_cleaned[col] = df_cleaned[col].astype(str).replace("NaT", "")

    df_cleaned = df_cleaned.where(pd.notnull(df_cleaned), None)
    joined_data = df_cleaned.to_dict(orient="records")
    print(len(joined_data), "rows prepared for chart_view.")

    # chart_meta stored earlier in session
    chart_meta = session.get("charts_meta", [])  # list of chart config dicts
    print(f"Loaded {len(chart_meta)} chart metadata entries from session.")

    # chart_files still used by your current UI if needed
    chart_files = session.get("chart_files", [])
    
    # build a map of unique values per column (strings) for the filter UI
    column_values = {}
    for col in df.columns:
        # dropna then unique and convert to python types (string or number)
        vals = df[col].dropna().unique().tolist()
        # convert numpy types to Python native
        vals = [v.item() if hasattr(v, "item") else v for v in vals]
        # sort strings, numbers works too
        try:
            vals_sorted = sorted(vals, key=lambda x: (str(type(x)), x))
        except Exception:
            vals_sorted = vals
        column_values[col] = vals_sorted

    # ensure static files ready (your existing waiting logic)
    if chart_files:
        static_paths = [os.path.join(BASE_DIR, "static", chart_file) for chart_file in chart_files]
        for _ in range(20):
            if all(os.path.exists(static_path) and os.path.getsize(static_path) > 1000 for static_path in static_paths):
                break
            time.sleep(0.1)

    return render_template(
        "chart_view.html",
        chart_files=chart_files,
        chart_meta=chart_meta,
        joined_data=joined_data,
        column_values=column_values
    )




@app.route("/table_view")
def table_view():
    joined_sheet = "joined_output"

    if not os.path.exists(EXCEL_FILE):
        flash("Excel file not found. Please upload again.")
        return redirect(url_for("data_model"))

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=joined_sheet)
    except Exception as e:
        flash(f"Error loading joined_output: {e}")
        return redirect(url_for("data_model"))
    
    # ----------- Replace NaN & NaT with empty string ----------
    df = df.replace({pd.NA: "", np.nan: "", "NaT": "", pd.NaT: ""})
    

    # ----------- Fix float columns ----------
    for col in df.columns:
        if df[col].dtype == 'float64':
            # If the column is float but should be int
            if df[col].dropna().apply(lambda x: float(x).is_integer() if pd.notnull(x) else True).all():
                df[col] = df[col].astype('Int64').astype(str).replace("<NA>", "")

    # ----------- Add Serial Number (S.No) ----------
    df.insert(0, "S.No", range(1, len(df)+1))

    # ----------- Format amount/currency columns ----------
    amount_keywords = ["amount", "total", "price", "cost", "transaction_amount"]
    currency_cols = [
        col for col in df.columns if any(keyword.lower() in col.lower() for keyword in amount_keywords)
    ]
    for col in currency_cols:
        def format_currency(x):
            try:
                # Try converting to float
                num = float(str(x).replace(",", "").replace("₹", "").strip())
                return f"₹{int(num):,}"
            except (ValueError, TypeError):
                # If it's empty, non-numeric, or invalid — return blank
                return ""
        df[col] = df[col].apply(format_currency)

    # ----------- Add totals row for amount columns ----------
    totals = {}
    for col in df.columns:
        if col in currency_cols:
            # Sum numeric values before formatting
            col_numeric_sum = pd.to_numeric(df[col].str.replace('₹','').str.replace(',',''), errors='coerce').sum()
            totals[col] = f"₹{int(col_numeric_sum):,}"
        elif col == "S.No":
            totals[col] = "Total"
        else:
            totals[col] = ""

    # Append total row
    df.loc[len(df)] = totals

    # --- Pagination ---
    page = int(request.args.get('page', 1))
    per_page = 20
    # Calculate total pages
    total_rows = len(df) - 1  # minus totals row
    total_pages = (total_rows + per_page - 1) // per_page
    # Convert to list of dicts for Jinja rendering
    data = df.to_dict(orient='records')
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_data = data[start_idx:end_idx]
    
    

    # Convert df to HTML table
    table_html = df.to_html(
        classes="table table-striped table-bordered table-sm text-center",
        index=False,
        escape=False
    )

    # # Apply CSS for total row bold + background
    # table_html = f"""
    # <style>
    #     table tbody tr:last-child {{
    #         font-weight: bold;
    #         background-color: #f0f0f0;
    #     }}
    #     table th, table td {{
    #         text-align: center;
    #     }}
    # </style>
    # {table_html}
    # """


    # CSS for total row if shown
    # style_html = ""
    # if show_totals:
    #     style_html = """
    #     <style>
    #         table tbody tr:last-child {
    #             font-weight: bold;
    #             background-color: #f0f0f0;
    #         }
    #         table th, table td {
    #             text-align: center;
    #         }
    #     </style>
    #     """
    # else:
    #     style_html = """
    #     <style>
    #         table th, table td {
    #             text-align: center;
    #         }
    #     </style>
    #     """

    

    return render_template("table_view.html", 
                           data=page_data,
                           columns=df.columns, 
                           page=page, 
                           total_pages=total_pages)


@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html', active_page='dashboard')

if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0")



